"""Service for generating reports in PDF and Excel formats."""

import logging
from typing import Dict, Optional
from datetime import date, datetime
from pathlib import Path
import pandas as pd
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.database.connection import get_db
from config.settings import settings

logger = logging.getLogger(__name__)


class RelatorioService:
    """Service for generating reports."""
    
    def __init__(self):
        """Initialize report service."""
        # Setup templates directory
        self.templates_dir = Path(__file__).parent.parent.parent / "templates" / "relatorios"
        self.templates_dir.mkdir(parents=True, exist_ok=True)
        
        self.jinja_env = Environment(
            loader=FileSystemLoader(str(self.templates_dir))
        )
        
        # Setup output directory
        self.output_dir = Path("/tmp/relatorios")
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def gerar_relatorio_diario_pdf(self, data: date) -> Optional[str]:
        """
        Generate daily report in PDF format.
        
        Args:
            data: Date for the report
            
        Returns:
            File path of generated PDF or None if error
        """
        try:
            db = get_db()
            cursor = db.cursor()
            
            # Get daily statistics
            cursor.execute("""
                SELECT 
                    COUNT(*) as total,
                    COALESCE(SUM(valor_total_estimado), 0) as valor_total,
                    COUNT(DISTINCT municipio_id) as municipios,
                    modalidade,
                    COUNT(*) as count_modalidade
                FROM licitacoes
                WHERE DATE(data_publicacao_pncp) = %s
                GROUP BY modalidade
                ORDER BY count_modalidade DESC
            """, (data,))
            
            modalidades = cursor.fetchall()
            
            # Get top licitacoes
            cursor.execute("""
                SELECT 
                    l.numero_compra,
                    m.nome as municipio,
                    l.objeto,
                    l.valor_total_estimado,
                    l.modalidade
                FROM licitacoes l
                JOIN municipios m ON l.municipio_id = m.id
                WHERE DATE(l.data_publicacao_pncp) = %s
                ORDER BY l.valor_total_estimado DESC
                LIMIT 10
            """, (data,))
            
            licitacoes = cursor.fetchall()
            cursor.close()
            
            # Calculate totals
            total_licitacoes = sum(m[4] for m in modalidades) if modalidades else 0
            valor_total = sum(m[1] for m in modalidades) if modalidades else 0
            
            # Prepare context
            context = {
                'data': data.strftime('%d/%m/%Y'),
                'total_licitacoes': total_licitacoes,
                'valor_total': valor_total,
                'modalidades': [
                    {
                        'nome': m[3],
                        'count': m[4]
                    }
                    for m in modalidades
                ],
                'licitacoes': [
                    {
                        'numero': l[0],
                        'municipio': l[1],
                        'objeto': l[2][:100] + '...' if len(l[2]) > 100 else l[2],
                        'valor': l[3],
                        'modalidade': l[4]
                    }
                    for l in licitacoes
                ],
                'generated_at': datetime.now().strftime('%d/%m/%Y %H:%M')
            }
            
            # Render HTML
            template = self.jinja_env.get_template('diario.html')
            html_content = template.render(context)
            
            # Generate PDF
            filename = f"relatorio_diario_{data.strftime('%Y%m%d')}.pdf"
            filepath = self.output_dir / filename
            
            HTML(string=html_content).write_pdf(str(filepath))
            
            logger.info(f"Daily report generated: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating daily report: {e}")
            return None
    
    def gerar_relatorio_licitacoes_excel(
        self,
        data_inicio: date,
        data_fim: date,
        municipio_id: Optional[int] = None
    ) -> Optional[str]:
        """
        Generate licitacoes report in Excel format.
        
        Args:
            data_inicio: Start date
            data_fim: End date
            municipio_id: Optional municipality filter
            
        Returns:
            File path of generated Excel or None if error
        """
        try:
            db = get_db()
            cursor = db.cursor()
            
            # Build query
            query = """
                SELECT 
                    l.numero_compra,
                    m.nome as municipio,
                    l.modalidade,
                    l.objeto,
                    l.valor_total_estimado,
                    l.data_abertura_proposta,
                    l.situacao,
                    l.data_publicacao_pncp
                FROM licitacoes l
                JOIN municipios m ON l.municipio_id = m.id
                WHERE DATE(l.data_publicacao_pncp) BETWEEN %s AND %s
            """
            params = [data_inicio, data_fim]
            
            if municipio_id:
                query += " AND l.municipio_id = %s"
                params.append(municipio_id)
            
            query += " ORDER BY l.data_publicacao_pncp DESC"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            cursor.close()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Licitações"
            
            # Headers
            headers = [
                'Número', 'Município', 'Modalidade', 'Objeto',
                'Valor Estimado', 'Abertura', 'Situação', 'Publicação'
            ]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row in rows:
                ws.append([
                    row[0],  # numero
                    row[1],  # municipio
                    row[2],  # modalidade
                    row[3][:100] if row[3] else '',  # objeto (truncated)
                    float(row[4]) if row[4] else 0,  # valor
                    row[5].strftime('%d/%m/%Y') if row[5] else '',  # abertura
                    row[6],  # situacao
                    row[7].strftime('%d/%m/%Y') if row[7] else ''  # publicacao
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12
            
            # Save file
            filename = f"licitacoes_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
            filepath = self.output_dir / filename
            
            wb.save(str(filepath))
            
            logger.info(f"Licitacoes Excel report generated: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return None
    
    def gerar_relatorio_fornecedores_excel(self) -> Optional[str]:
        """
        Generate suppliers ranking report in Excel format.
        
        Returns:
            File path of generated Excel or None if error
        """
        try:
            db = get_db()
            cursor = db.cursor()
            
            # Get fornecedores statistics
            cursor.execute("""
                SELECT 
                    f.razao_social,
                    f.cnpj,
                    f.porte,
                    COUNT(r.id) as total_vitorias,
                    COALESCE(SUM(r.valor_proposta), 0) as valor_total_ganho,
                    COUNT(DISTINCT r.licitacao_id) as licitacoes_participadas
                FROM fornecedores f
                LEFT JOIN resultados r ON r.fornecedor_id = f.id
                WHERE r.vencedor = true
                GROUP BY f.id, f.razao_social, f.cnpj, f.porte
                ORDER BY valor_total_ganho DESC
                LIMIT 100
            """)
            
            rows = cursor.fetchall()
            cursor.close()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Ranking Fornecedores"
            
            # Headers
            headers = [
                'Ranking', 'Razão Social', 'CNPJ', 'Porte',
                'Vitórias', 'Valor Total Ganho', 'Participações'
            ]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for i, row in enumerate(rows, start=1):
                ws.append([
                    i,  # ranking
                    row[0],  # razao_social
                    row[1],  # cnpj
                    row[2],  # porte
                    row[3],  # vitorias
                    float(row[4]),  # valor_total
                    row[5]  # participacoes
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15
            
            # Save file
            filename = f"fornecedores_{datetime.now().strftime('%Y%m%d')}.xlsx"
            filepath = self.output_dir / filename
            
            wb.save(str(filepath))
            
            logger.info(f"Fornecedores Excel report generated: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating fornecedores report: {e}")
            return None


# Global instance
relatorio_service = RelatorioService()
